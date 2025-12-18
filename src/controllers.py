import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from src.database import Database
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class FinanceiroController:
    def __init__(self):
        self.db = Database()

    def adicionar_despesa(self, data: str, tipo: str, categoria: str, descricao: str, valor: float, recorrencia_meses: int = 0):
        """Adiciona uma despesa. Aceita data em 'DD/MM/YYYY' ou 'MM/YY'. Se recorrência > 0, cria automaticamente para os próximos meses.
        Armazena data no formato 'DD/MM/YYYY' para compatibilidade com o restante da aplicação."""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()

            # Tentar interpretar data em vários formatos
            parse_formats = ["%d/%m/%Y", "%m/%y", "%Y-%m-%d", "%d-%m-%Y"]
            data_dt = None
            for fmt in parse_formats:
                try:
                    data_dt = datetime.strptime(data, fmt)
                    # Se veio apenas mês/ano (%m/%y), definir dia como 1
                    if fmt == "%m/%y":
                        data_dt = data_dt.replace(day=1)
                    break
                except Exception:
                    continue

            if data_dt is None:
                raise ValueError("Formato de data inválido. Use DD/MM/YYYY ou MM/YY.")

            # Determinar quantas vezes repetir
            if tipo == "Fixa":
                repeticoes = 12  # Fixas se repetem por 1 ano
            elif recorrencia_meses and recorrencia_meses > 0:
                repeticoes = recorrencia_meses
            else:
                repeticoes = 1  # Apenas uma vez (Variável sem recorrência)

            # Inserir a despesa original e as recorrentes (armazena como DD/MM/YYYY)
            for i in range(repeticoes):
                data_futura = data_dt + relativedelta(months=i)
                data_str = data_futura.strftime("%d/%m/%Y")

                cursor.execute(
                    "INSERT INTO despesas (data, tipo, categoria, descricao, valor, recorrencia_meses) VALUES (?, ?, ?, ?, ?, ?)",
                    (data_str, tipo, categoria, descricao, valor, recorrencia_meses)
                )

            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Erro ao adicionar despesa: {e}")
            raise e

    def excluir_despesa(self, despesa_id: int):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM despesas WHERE id = ?", (despesa_id,))
        conn.commit()
        conn.close()

    def buscar_despesas_mes(self, mes: int, ano: int) -> pd.DataFrame:
        """Retorna DataFrame com despesas do mês/ano selecionado."""
        conn = self.db.get_connection()
        # Formato da data no banco é YYYY-MM-DD ou DD/MM/YYYY. 
        # Aqui assumimos input padronizado. Filtramos via SQL é mais eficiente.
        # SQLite não tem tipo DATE nativo forte, faremos filtro no Pandas para flexibilidade
        # ou construímos a query com LIKE. Pela robustez, vamos trazer tudo e filtrar no pandas (escala pequena)
        # ou usar strftime do sqlite.
        
        query = "SELECT * FROM despesas"
        df = pd.read_sql_query(query, conn)
        conn.close()

        if df.empty:
            return df

        # Conversão segura de datas
        df['data_dt'] = pd.to_datetime(df['data'], format='%d/%m/%Y', errors='coerce')
        df = df[
            (df['data_dt'].dt.month == mes) & 
            (df['data_dt'].dt.year == ano)
        ]
        return df.sort_values(by='data_dt')

    def get_configuracoes(self):
        """Retorna (salario_1, salario_2)"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT salario_1, salario_2 FROM configuracoes WHERE id = 1")
        resultado = cursor.fetchone()
        conn.close()
        return resultado if resultado else (0.0, 0.0)

    def salvar_configuracoes(self, sal1: float, sal2: float):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE configuracoes SET salario_1 = ?, salario_2 = ? WHERE id = 1", (sal1, sal2))
        conn.commit()
        conn.close()
    
    def adicionar_receita_extra(self, mes: int, ano: int, descricao: str, valor: float):
        """Adiciona uma receita extra para um mês/ano específico"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT OR REPLACE INTO receitas_extras (mes, ano, descricao, valor) VALUES (?, ?, ?, ?)",
            (mes, ano, descricao, valor)
        )
        conn.commit()
        conn.close()
    
    def buscar_receitas_extras_mes(self, mes: int, ano: int):
        """Busca receitas extras de um mês/ano"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM receitas_extras WHERE mes = ? AND ano = ?", (mes, ano))
        resultado = cursor.fetchall()
        conn.close()
        return resultado
    
    def adicionar_categoria(self, nome: str, icone: str):
        """Adiciona uma nova categoria personalizada."""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO categorias (nome, icone) VALUES (?, ?)", (nome, icone))
            conn.commit()
            conn.close()
        except sqlite3.IntegrityError:
            raise Exception("Categoria já existe!")
        except Exception as e:
            raise e
    
    def buscar_categorias(self):
        """Retorna todas as categorias com seus ícones."""
        conn = self.db.get_connection()
        query = "SELECT nome, icone FROM categorias ORDER BY nome"
        df = pd.read_sql_query(query, conn)
        conn.close()
        return df
    
    def excluir_categoria(self, nome: str):
        """Exclui uma categoria personalizada."""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM categorias WHERE nome = ?", (nome,))
        conn.commit()
        conn.close()
    
    def excluir_receita_extra(self, receita_id: int):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM receitas_extras WHERE id = ?", (receita_id,))
        conn.commit()
        conn.close()
    
    def calcular_totais_mes(self, mes: int, ano: int):
        """Retorna (receita_total, despesas_total, saldo)"""
        salarios = self.get_configuracoes()
        receita_total = sum(salarios)
        
        # Adicionar receitas extras do mês
        receitas_extras = self.buscar_receitas_extras_mes(mes, ano)
        total_extras = sum([r[4] for r in receitas_extras])  # r[4] é o valor
        receita_total += total_extras
        
        df = self.buscar_despesas_mes(mes, ano)
        despesas_total = df['valor'].sum() if not df.empty else 0.0
        
        saldo = receita_total - despesas_total
        
        return (receita_total, despesas_total, saldo)

    def exportar_relatorio(self, mes: int, ano: int, caminho_arquivo: str):
        """Gera Excel estilizado e profissional."""
        df = self.buscar_despesas_mes(mes, ano)
        salarios = self.get_configuracoes()
        receita_total = sum(salarios)
        total_despesas = df['valor'].sum() if not df.empty else 0.0
        saldo = receita_total - total_despesas

        wb = Workbook()
        ws = wb.active
        ws.title = f"{mes:02d}-{ano}"

        # Cabeçalhos
        headers = ["ID", "Data", "Tipo", "Categoria", "Descrição", "Valor (R$)"]
        ws.append(headers)

        # Estilo Cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for _, row in df.iterrows():
            ws.append([row['id'], row['data'], row['tipo'], row['categoria'], row['descricao'], row['valor']])

        # Formatação Monetária e Auto-fit
        for row in ws.iter_rows(min_row=2, max_col=6):
            # Coluna de valor é a 6 agora
            row[5].number_format = 'R$ #,##0.00'

        # Resumo no final
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=5, value="RECEITA TOTAL:").font = Font(bold=True)
        ws.cell(row=last_row, column=6, value=receita_total).number_format = 'R$ #,##0.00'
        
        ws.cell(row=last_row+1, column=5, value="TOTAL DESPESAS:").font = Font(bold=True, color="FF0000")
        ws.cell(row=last_row+1, column=6, value=total_despesas).number_format = 'R$ #,##0.00'

        ws.cell(row=last_row+2, column=5, value="SALDO FINAL:").font = Font(bold=True)
        saldo_cell = ws.cell(row=last_row+2, column=6, value=saldo)
        saldo_cell.number_format = 'R$ #,##0.00'
        
        # Cor condicional no Excel
        if saldo >= 0:
            saldo_cell.font = Font(color="006100", bold=True) # Verde
        else:
            saldo_cell.font = Font(color="9C0006", bold=True) # Vermelho

        # Ajuste de largura das colunas
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save(caminho_arquivo)
        print(f"Relatório salvo em: {caminho_arquivo}")
